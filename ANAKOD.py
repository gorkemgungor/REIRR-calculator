import sys
import os
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QLabel, QLineEdit, QPushButton, QTableWidget,
    QTableWidgetItem, QGroupBox, QFileDialog,
    QMessageBox, QHeaderView, QTabWidget, QDateEdit, QComboBox, QCheckBox
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QFont, QColor, QBrush, QPixmap
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

OAR_DATA = [
    {"name": "Bladder",                          "ab": 2.5, "limit": 85,   "reduction": [0, 10, 25, 50]},
    {"name": "Brachial Plexus",                  "ab": 2.5, "limit": 70,   "reduction": [0, 10, 25, 50]},
    {"name": "Brainstem",                        "ab": 2.5, "limit": 64,   "reduction": [0, 10, 25, 50]},
    {"name": "Cauda Equina",                     "ab": 2.5, "limit": 60,   "reduction": [0, 10, 25, 50]},
    {"name": "Chest Wall",                       "ab": 2.5, "limit": 100,  "reduction": [0, 10, 25, 50]},
    {"name": "Colon",                            "ab": 2.5, "limit": 70,   "reduction": [0, 10, 25, 50]},
    {"name": "Duodenum",                         "ab": 2.5, "limit": 54,   "reduction": [0,  0, 10, 25]},
    {"name": "Esophagus",                        "ab": 2.5, "limit": 70,   "reduction": [0, 10, 25, 50]},
    {"name": "Great Vessels",                    "ab": 2.5, "limit": 100,  "reduction": [0, 10, 25, 50]},
    {"name": "Heart",                            "ab": 2.5, "limit": 70,   "reduction": [0, 10, 25, 50]},
    {"name": "Kidneys",                          "ab": 2.5, "limit": None,  "reduction": [0,  0,  0,  0]},
    {"name": "Optic Chiasm",                     "ab": 2.5, "limit": 54,   "reduction": [0, 10, 25, 50]},
    {"name": "Optic Nerve",                      "ab": 2.5, "limit": 54,   "reduction": [0, 10, 25, 50]},
    {"name": "Rectum",                           "ab": 2.5, "limit": 80,   "reduction": [0, 10, 25, 50]},
    {"name": "Retina",                           "ab": 2.5, "limit": 50,   "reduction": [0, 10, 25, 50]},
    {"name": "Sacral Plexus",                    "ab": 2.5, "limit": 70,   "reduction": [0, 10, 25, 50]},
    {"name": "Small Bowel",                      "ab": 2.5, "limit": 54,   "reduction": [0,  0, 25, 25]},
    {"name": "Spinal Cord",                      "ab": 2.5, "limit": 50,   "reduction": [0, 10, 25, 50]},
    {"name": "Spinal Cord (<2mm from target)",   "ab": 2.5, "limit": 55,   "reduction": [0, 10, 25, 50]},
    {"name": "Stomach",                          "ab": 2.5, "limit": 54,   "reduction": [0,  0, 25, 25]},
    {"name": "Trachea/Bronchus",                 "ab": 2.5, "limit": 70,   "reduction": [0, 10, 25, 50]},
    {"name": "Liver",                            "ab": 2.5, "limit": None,  "reduction": [0,  0, 50, 100]},
    {"name": "Lungs",                            "ab": 2.5, "limit": None,  "reduction": [0,  0, 25, 50]},
]

TIME_INTERVALS = ["< 3 mo", "3-6 mo", "6 mo - 1 yr", "1 yr - 3 yrs"]

COL_TICK   = 0
COL_OAR    = 1
COL_AB     = 2
COL_LIMIT  = 3
COL_RED    = 4
COL_RED2   = 5
COL_ALLOW  = 6
COL_C1     = 7
COL_C2     = 8
COL_C3     = 9
COL_SUM    = 10
COL_STATUS = 11

INPUT_BG_DEFAULT  = "background-color: white;"
INPUT_BG_SELECTED = "background-color: #b4dcff;"
INPUT_BG_FROZEN   = "background-color: #d2ebff;"
INPUT_BG_DIM      = "background-color: #dcdcdc;"


def calc_eqd2(dose_per_fr, n_fractions, ab):
    if dose_per_fr <= 0 or n_fractions <= 0:
        return 0.0
    total = dose_per_fr * n_fractions
    return total * (dose_per_fr + ab) / (2 + ab)


def calc_bed(dose_per_fr, n_fractions, ab):
    if dose_per_fr <= 0 or n_fractions <= 0:
        return 0.0
    total = dose_per_fr * n_fractions
    return total * (1 + dose_per_fr / ab)


class CourseWidget(QGroupBox):
    def __init__(self, label, on_change=None, parent=None):
        super().__init__(label, parent)
        self.setFont(QFont("Arial", 13, QFont.Weight.Light))
        self.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        self._updating = False
        self._on_change = on_change

        layout = QGridLayout()
        layout.setSpacing(6)
        layout.setContentsMargins(12, 12, 12, 12)

        def lbl(text):
            l = QLabel(text)
            l.setFont(QFont("Arial", 13, QFont.Weight.Light))
            return l

        def out(text="0.00"):
            l = QLabel(text)
            l.setFont(QFont("Arial", 13, QFont.Weight.Bold))
            l.setAlignment(Qt.AlignmentFlag.AlignRight)
            return l

        layout.addWidget(lbl("Dose/Fr. (Gy):"), 0, 0)
        self.dose_fr = QLineEdit("0")
        self.dose_fr.setMinimumWidth(90)
        self.dose_fr.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.dose_fr, 0, 1)

        layout.addWidget(lbl("# of Fractions:"), 1, 0)
        self.n_fr = QLineEdit("0")
        self.n_fr.setMinimumWidth(90)
        self.n_fr.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.n_fr, 1, 1)

        layout.addWidget(lbl("Total Dose (Gy):"), 2, 0)
        self.total_dose = QLineEdit("0.00")
        self.total_dose.setMinimumWidth(90)
        self.total_dose.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.total_dose, 2, 1)

        layout.addWidget(lbl("EQD2 (α/β=3):"), 3, 0)
        self.eqd2_3 = out(); layout.addWidget(self.eqd2_3, 3, 1)

        layout.addWidget(lbl("BED (α/β=3):"), 4, 0)
        self.bed_3 = out(); layout.addWidget(self.bed_3, 4, 1)

        layout.addWidget(lbl("EQD2 (α/β=10):"), 5, 0)
        self.eqd2_10 = out(); layout.addWidget(self.eqd2_10, 5, 1)

        layout.addWidget(lbl("BED (α/β=10):"), 6, 0)
        self.bed_10 = out(); layout.addWidget(self.bed_10, 6, 1)

        layout.addWidget(lbl("Manual α/β:"), 7, 0)
        self.manual_ab = QComboBox()
        self.manual_ab.addItems([f"{v/10:.1f}" for v in range(5, 105, 5)])
        self.manual_ab.setCurrentText("2.5")
        self.manual_ab.setMinimumWidth(90)
        self.manual_ab.setEditable(True)
        self.manual_ab.lineEdit().setReadOnly(True)
        self.manual_ab.lineEdit().setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.manual_ab, 7, 1)

        layout.addWidget(lbl("EQD2 (manual):"), 8, 0)
        self.eqd2_man = out(); layout.addWidget(self.eqd2_man, 8, 1)

        layout.addWidget(lbl("BED (manual):"), 9, 0)
        self.bed_man = out(); layout.addWidget(self.bed_man, 9, 1)

        self.setLayout(layout)

        self.dose_fr.textChanged.connect(self._dose_fr_changed)
        self.n_fr.textChanged.connect(self._n_fr_changed)
        self.total_dose.textChanged.connect(self._total_changed)
        self.manual_ab.currentTextChanged.connect(self._ab_changed)

    # ── 3-way calculation ──────────────────────────────────────
    def _dose_fr_changed(self):
        """dose_fr changed → compute total = dose_fr x n_fr"""
        if self._updating:
            return
        self._updating = True
        d, n, ab = self.get_values()
        self.total_dose.setText(f"{d * n:.2f}")
        self._recalc(d, n, ab)
        self._updating = False
        if self._on_change:
            self._on_change()

    def _n_fr_changed(self):
        """n_fr changed:
           - if dose_fr > 0 → total = dose_fr x n_fr
           - if dose_fr == 0 and total > 0 → dose_fr = total / n_fr
        """
        if self._updating:
            return
        self._updating = True
        try:
            d = float(self.dose_fr.text())
        except ValueError:
            d = 0
        try:
            n = float(self.n_fr.text())
        except ValueError:
            n = 0
        try:
            total = float(self.total_dose.text())
        except ValueError:
            total = 0
        try:
            ab = float(self.manual_ab.currentText())
        except ValueError:
            ab = 2.5

        if d > 0:
            total = d * n
            self.total_dose.setText(f"{total:.2f}")
        elif n > 0 and total > 0:
            d = total / n
            self.dose_fr.setText(f"{d:.2f}")
        self._recalc(d, n, ab)
        self._updating = False
        if self._on_change:
            self._on_change()

    def _total_changed(self):
        """total changed:
           - if dose_fr > 0 → n_fr = total / dose_fr
           - elif n_fr > 0  → dose_fr = total / n_fr
        """
        if self._updating:
            return
        self._updating = True
        try:
            total = float(self.total_dose.text())
        except ValueError:
            total = 0
        try:
            d = float(self.dose_fr.text())
        except ValueError:
            d = 0
        try:
            n = float(self.n_fr.text())
        except ValueError:
            n = 0
        try:
            ab = float(self.manual_ab.currentText())
        except ValueError:
            ab = 2.5

        if d > 0:
            n = total / d
            self.n_fr.setText(f"{n:.2f}")
        elif n > 0:
            d = total / n
            self.dose_fr.setText(f"{d:.2f}")
        self._recalc(d, n, ab)
        self._updating = False
        if self._on_change:
            self._on_change()

    def _ab_changed(self):
        if self._updating:
            return
        self._updating = True
        d, n, ab = self.get_values()
        self._recalc(d, n, ab)
        self._updating = False
        if self._on_change:
            self._on_change()

    def _recalc(self, d, n, ab):
        self.eqd2_3.setText(f"{calc_eqd2(d, n, 3):.2f}")
        self.bed_3.setText(f"{calc_bed(d, n, 3):.2f}")
        self.eqd2_10.setText(f"{calc_eqd2(d, n, 10):.2f}")
        self.bed_10.setText(f"{calc_bed(d, n, 10):.2f}")
        self.eqd2_man.setText(f"{calc_eqd2(d, n, ab):.2f}")
        self.bed_man.setText(f"{calc_bed(d, n, ab):.2f}")

    def get_values(self):
        try:
            d = float(self.dose_fr.text())
        except ValueError:
            d = 0
        try:
            n = float(self.n_fr.text())
        except ValueError:
            n = 0
        try:
            ab = float(self.manual_ab.currentText())
        except ValueError:
            ab = 2.5
        return d, n, ab

    def has_values(self):
        d, n, _ = self.get_values()
        return d > 0 or n > 0

    def get_eqd2_manual(self):
        d, n, ab = self.get_values()
        return calc_eqd2(d, n, ab)

    def set_inputs_enabled(self, enabled):
        for w in [self.dose_fr, self.n_fr, self.total_dose, self.manual_ab]:
            w.setEnabled(enabled)

    def clear(self):
        self._updating = True
        self.dose_fr.setText("0")
        self.n_fr.setText("0")
        self.total_dose.setText("0.00")
        self.manual_ab.setCurrentText("2.5")
        self._recalc(0, 0, 2.5)
        self._updating = False


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("LQ Model Re-Irradiation Calculator - Radiotherapy Dose Calculator")
        self.setMinimumSize(1300, 900)
        self.current_oar_idx = None
        self.frozen_oars = set()
        self.oar_raw = [[0.0, 0.0, 0.0] for _ in OAR_DATA]

        en_label = QLabel()
        en_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "EN.png")
        en_pixmap = QPixmap(en_path)
        if not en_pixmap.isNull():
            en_label.setPixmap(en_pixmap.scaledToHeight(60, Qt.TransformationMode.SmoothTransformation))
        en_label.setAlignment(Qt.AlignmentFlag.AlignLeft)
        en_label.setContentsMargins(12, 6, 0, 6)

        logo_label = QLabel()
        logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "icon_source.png")
        pixmap = QPixmap(logo_path)
        if not pixmap.isNull():
            logo_label.setPixmap(pixmap.scaledToHeight(60, Qt.TransformationMode.SmoothTransformation))
        logo_label.setAlignment(Qt.AlignmentFlag.AlignRight)
        logo_label.setContentsMargins(0, 6, 12, 6)

        credits_widget = QLabel("Coded by Görkem Güngör & Developed by Artunç Türe")
        credits_widget.setAlignment(Qt.AlignmentFlag.AlignCenter)
        credits_widget.setStyleSheet("color: black; font-size: 11px;")

        header_bar = QWidget()
        hl = QHBoxLayout(header_bar)
        hl.setContentsMargins(0, 0, 0, 0)
        hl.addWidget(en_label)
        hl.addStretch()
        hl.addWidget(credits_widget)
        hl.addStretch()
        hl.addWidget(logo_label)

        tabs = QTabWidget()
        tabs.addTab(self.build_patient_tab(), "Patient Info")
        tabs.addTab(self.build_main_tab(), "Re-Irradiation EQD2 Calculator")

        container = QWidget()
        cl = QVBoxLayout()
        cl.setContentsMargins(0, 0, 0, 0)
        cl.setSpacing(0)
        cl.addWidget(header_bar)
        cl.addWidget(tabs)
        container.setLayout(cl)
        self.setCentralWidget(container)

    # ── Patient tab ───────────────────────────────────────────
    def build_patient_tab(self):
        outer = QWidget()
        ol = QHBoxLayout()
        ol.addStretch()
        w = QWidget(); w.setFixedWidth(420)
        layout = QGridLayout(); layout.setSpacing(10)
        fields = [
            ("Patient Name:", "patient_name", None),
            ("Patient ID:", "patient_id", None),
            ("C1 Plan Date:", "dob", "date"),
            ("C2 Plan Date:", "plan_date", "date"),
            ("Medical Physicist:", "diagnosis", None),
            ("Physician:", "physician", None),
            ("Notes:", "notes", None),
        ]
        for i, (label, attr, ftype) in enumerate(fields):
            layout.addWidget(QLabel(label), i, 0)
            if ftype == "date":
                widget = QDateEdit(); widget.setCalendarPopup(True); widget.setDate(QDate.currentDate())
                widget.setAlignment(Qt.AlignmentFlag.AlignCenter)
            else:
                widget = QLineEdit()
            setattr(self, attr, widget)
            layout.addWidget(widget, i, 1)
        layout.setRowStretch(len(fields), 1)
        w.setLayout(layout)
        ol.addWidget(w); ol.addStretch()
        outer.setLayout(ol)
        return outer

    # ── Calculator tab ────────────────────────────────────────
    def build_main_tab(self):
        widget = QWidget()
        main_layout = QVBoxLayout()
        main_layout.setSpacing(8)

        # ── OAR selector ──────────────────────────────────────
        oar_sel_layout = QHBoxLayout()
        oar_lbl = QLabel("OAR:")
        oar_lbl.setFont(QFont("Arial", 15, QFont.Weight.Bold))
        oar_sel_layout.addWidget(oar_lbl)
        self.oar_selector = QComboBox()
        self.oar_selector.setMinimumWidth(300)
        f = QFont("Arial")
        f.setPointSizeF(13.5)
        self.oar_selector.setFont(f)
        self.oar_selector.setEditable(True)
        self.oar_selector.lineEdit().setReadOnly(True)
        self.oar_selector.lineEdit().setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.oar_selector.addItem("— Select OAR —")
        for oar in OAR_DATA:
            self.oar_selector.addItem(oar["name"])
        self.oar_selector.currentIndexChanged.connect(self.on_oar_selected)
        oar_sel_layout.addWidget(self.oar_selector)
        oar_sel_layout.addStretch()
        main_layout.addLayout(oar_sel_layout)

        # ── Course widgets ────────────────────────────────────
        course_layout = QGridLayout()
        course_layout.setSpacing(6)
        for col in range(4):
            course_layout.setColumnStretch(col, 1)
        self.c1 = CourseWidget("COURSE 1", on_change=self.on_course_changed)
        self.c2 = CourseWidget("COURSE 2", on_change=self.on_course_changed)
        self.c3 = CourseWidget("COURSE 3", on_change=self.on_course_changed)

        sum_box = QGroupBox("SUM")
        sum_box.setFont(QFont("Arial", 15, QFont.Weight.Bold))
        sum_box.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        sl = QGridLayout(); sl.setSpacing(6); sl.setContentsMargins(12, 12, 12, 12)

        def s_lbl(t):
            l = QLabel(t); l.setFont(QFont("Arial", 15, QFont.Weight.Light)); return l

        def s_out():
            l = QLabel("0.00"); l.setFont(QFont("Arial", 15, QFont.Weight.Bold))
            l.setAlignment(Qt.AlignmentFlag.AlignRight); return l

        self.sum_total    = s_out(); sl.addWidget(s_lbl("Total Dose (Gy):"), 0, 0); sl.addWidget(self.sum_total, 0, 1)
        self.sum_eqd2_3   = s_out(); sl.addWidget(s_lbl("EQD2 (α/β=3):"),   1, 0); sl.addWidget(self.sum_eqd2_3, 1, 1)
        self.sum_bed_3    = s_out(); sl.addWidget(s_lbl("BED (α/β=3):"),     2, 0); sl.addWidget(self.sum_bed_3, 2, 1)
        self.sum_eqd2_10  = s_out(); sl.addWidget(s_lbl("EQD2 (α/β=10):"),  3, 0); sl.addWidget(self.sum_eqd2_10, 3, 1)
        self.sum_bed_10   = s_out(); sl.addWidget(s_lbl("BED (α/β=10):"),   4, 0); sl.addWidget(self.sum_bed_10, 4, 1)
        self.sum_eqd2_man = s_out(); sl.addWidget(s_lbl("EQD2 (manual):"),  5, 0); sl.addWidget(self.sum_eqd2_man, 5, 1)
        self.sum_bed_man  = s_out(); sl.addWidget(s_lbl("BED (manual):"),   6, 0); sl.addWidget(self.sum_bed_man, 6, 1)
        sum_box.setLayout(sl)

        # ── Reduction controls (below their respective courses) ─
        # C1 reduction
        self.c1_red_check = QCheckBox("C1 Time Reduction:")
        self.c1_red_check.setFont(QFont("Arial", 11))
        self.c1_red_check.setChecked(False)
        self.c1_time_combo = QComboBox()
        self.c1_time_combo.addItems(TIME_INTERVALS)
        self.c1_time_combo.setEnabled(False)
        self.c1_time_combo.setEditable(True)
        self.c1_time_combo.lineEdit().setReadOnly(True)
        self.c1_time_combo.lineEdit().setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.c1_red_check.stateChanged.connect(
            lambda s: self.c1_time_combo.setEnabled(s == Qt.CheckState.Checked.value))
        self.c1_red_check.stateChanged.connect(self.update_oar_table)
        self.c1_time_combo.currentIndexChanged.connect(self.update_oar_table)

        c1_red_layout = QHBoxLayout()
        c1_red_layout.setContentsMargins(0, 2, 0, 0)
        c1_red_layout.addStretch()
        c1_red_layout.addWidget(self.c1_red_check)
        c1_red_layout.addWidget(self.c1_time_combo)
        c1_red_layout.addStretch()

        # C2 reduction (enabled only when C3 has values)
        self.c2_red_check = QCheckBox("C2 Time Reduction:")
        self.c2_red_check.setFont(QFont("Arial", 11))
        self.c2_red_check.setChecked(False)
        self.c2_red_check.setEnabled(False)
        self.c2_time_combo = QComboBox()
        self.c2_time_combo.addItems(TIME_INTERVALS)
        self.c2_time_combo.setEnabled(False)
        self.c2_time_combo.setEditable(True)
        self.c2_time_combo.lineEdit().setReadOnly(True)
        self.c2_time_combo.lineEdit().setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.c2_red_check.stateChanged.connect(
            lambda s: self.c2_time_combo.setEnabled(
                s == Qt.CheckState.Checked.value and self.c3.has_values()))
        self.c2_red_check.stateChanged.connect(self.update_oar_table)
        self.c2_time_combo.currentIndexChanged.connect(self.update_oar_table)

        c2_red_layout = QHBoxLayout()
        c2_red_layout.setContentsMargins(0, 2, 0, 0)
        c2_red_layout.addStretch()
        c2_red_layout.addWidget(self.c2_red_check)
        c2_red_layout.addWidget(self.c2_time_combo)
        c2_red_layout.addStretch()

        # Row 0: all 4 GroupBoxes → equal width & height
        course_layout.addWidget(self.c1,  0, 0)
        course_layout.addWidget(self.c2,  0, 1)
        course_layout.addWidget(self.c3,  0, 2)
        course_layout.addWidget(sum_box,  0, 3)
        # Row 1: reduction controls only under C1 and C2
        course_layout.addLayout(c1_red_layout, 1, 0, Qt.AlignmentFlag.AlignCenter)
        course_layout.addLayout(c2_red_layout, 1, 1, Qt.AlignmentFlag.AlignCenter)
        main_layout.addLayout(course_layout)

        # Disable courses until OAR selected
        self._set_courses_enabled(False)

        # ── OAR Table ─────────────────────────────────────────
        oar_group = QGroupBox("Organs At Risk (OAR) Dose Limits")
        oar_vl = QVBoxLayout()
        self.oar_table = QTableWidget()
        headers = ["✓", "OAR", "α/β\n(Gy)", "Dose Limit\n(EQD2 Gy)",
                   "C1\nReduction%", "C2\nReduction%", "Allowed\nEQD2 (Gy)",
                   "C1 Plan\nEQD2", "C2 Plan\nEQD2", "C3 Plan\nEQD2",
                   "SUM PLAN\nEQD2", "Status"]
        self.oar_table.setColumnCount(len(headers))
        self.oar_table.setHorizontalHeaderLabels(headers)
        self.oar_table.setRowCount(len(OAR_DATA))
        hdr = self.oar_table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(COL_TICK, QHeaderView.ResizeMode.Fixed)
        self.oar_table.setColumnWidth(COL_TICK, 44)
        hdr.setSectionResizeMode(COL_OAR, QHeaderView.ResizeMode.Interactive)
        self.oar_table.setColumnWidth(COL_OAR, 200)
        hdr.setStretchLastSection(True)
        self.oar_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.oar_table.setMinimumHeight(380)

        self.oar_ticks = []

        for row, oar in enumerate(OAR_DATA):
            tick = QCheckBox()
            tick.setEnabled(False)
            tick.stateChanged.connect(lambda state, r=row: self.toggle_freeze(r, state))
            tick_w = QWidget()
            tl = QHBoxLayout(tick_w)
            tl.setContentsMargins(0, 0, 0, 0)
            tl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            tl.addWidget(tick)
            self.oar_table.setCellWidget(row, COL_TICK, tick_w)
            self.oar_ticks.append(tick)

            limit_str = f"{oar['limit']}" if oar['limit'] is not None else "—"
            for col, val in [(COL_OAR, oar["name"]), (COL_AB, str(oar["ab"])),
                             (COL_LIMIT, limit_str), (COL_RED, ""), (COL_RED2, ""),
                             (COL_ALLOW, ""), (COL_SUM, ""), (COL_STATUS, "")]:
                it = QTableWidgetItem(val)
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.oar_table.setItem(row, col, it)

            for col in [COL_C1, COL_C2, COL_C3]:
                item = QTableWidgetItem("0")
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.oar_table.setItem(row, col, item)

        oar_vl.addWidget(self.oar_table)
        oar_group.setLayout(oar_vl)
        main_layout.addWidget(oar_group)

        # ── Buttons ───────────────────────────────────────────
        btn_layout = QHBoxLayout()
        for label, slot in [("Clear", self.clear_all),
                             ("Export PDF", self.export_pdf),
                             ("Export Excel", self.export_excel)]:
            btn = QPushButton(label)
            btn.clicked.connect(slot)
            btn.setFixedHeight(34)
            btn_layout.addWidget(btn)
        btn_layout.addStretch()
        main_layout.addLayout(btn_layout)

        widget.setLayout(main_layout)
        self.update_sum()
        self.update_oar_table()
        return widget

    # ── Helpers ───────────────────────────────────────────────
    def _set_courses_enabled(self, enabled):
        for c in [self.c1, self.c2, self.c3]:
            c.set_inputs_enabled(enabled)

    def _set_ticks_enabled(self, enabled):
        for tick in self.oar_ticks:
            tick.setEnabled(enabled)

    def _update_c2_reduction_enabled(self):
        has_c3 = self.c3.has_values()
        self.c2_red_check.setEnabled(has_c3)
        if not has_c3:
            self.c2_red_check.setChecked(False)
            self.c2_time_combo.setEnabled(False)

    # ── OAR selector ──────────────────────────────────────────
    def on_oar_selected(self, combo_idx):
        if combo_idx == 0:
            self.current_oar_idx = None
            self._set_courses_enabled(False)
            self._set_ticks_enabled(False)
            self.apply_row_highlight()
            return
        self.current_oar_idx = combo_idx - 1
        self._set_courses_enabled(True)
        self._set_ticks_enabled(True)
        self._autofill_oar(self.current_oar_idx)
        self.apply_row_highlight()

    def _autofill_oar(self, oar_idx):
        if oar_idx in self.frozen_oars:
            return
        for col_idx, course in enumerate([self.c1, self.c2, self.c3]):
            d, n, ab = course.get_values()
            self.oar_raw[oar_idx][col_idx] = calc_eqd2(d, n, ab)
        # Update α/β column to reflect current manual α/β (use c1 as reference)
        _, _, ab_display = self.c1.get_values()
        ab_item = self.oar_table.item(oar_idx, COL_AB)
        if ab_item:
            ab_item.setText(str(ab_display))
        self.update_oar_table()

    def on_course_changed(self):
        self.update_sum()
        self._update_c2_reduction_enabled()
        if self.current_oar_idx is not None and self.current_oar_idx not in self.frozen_oars:
            self._autofill_oar(self.current_oar_idx)
        else:
            self.update_oar_table()

    def toggle_freeze(self, row, state):
        frozen = (state == Qt.CheckState.Checked.value)
        if frozen:
            self.frozen_oars.add(row)
        else:
            self.frozen_oars.discard(row)
        self.apply_row_highlight()

    # ── Row highlight ─────────────────────────────────────────
    def apply_row_highlight(self):
        COLOR_SELECTED = QColor(180, 220, 255)
        COLOR_FROZEN   = QColor(210, 235, 255)
        COLOR_DIM      = QColor(220, 220, 220)
        COLOR_DEFAULT  = QColor(255, 255, 255)

        any_selected = self.current_oar_idx is not None

        for row in range(len(OAR_DATA)):
            if row == self.current_oar_idx:
                bg = COLOR_SELECTED
                inp_ss = INPUT_BG_SELECTED
            elif row in self.frozen_oars:
                bg = COLOR_FROZEN
                inp_ss = INPUT_BG_FROZEN
            elif any_selected:
                bg = COLOR_DIM
                inp_ss = INPUT_BG_DIM
            else:
                bg = COLOR_DEFAULT
                inp_ss = INPUT_BG_DEFAULT

            # Color all item cells except Status and Allowed (always blue)
            for col in range(COL_STATUS):
                if col == COL_ALLOW:
                    continue
                item = self.oar_table.item(row, col)
                if item:
                    item.setBackground(QBrush(bg))

    # ── Sum ───────────────────────────────────────────────────
    def update_sum(self):
        totals = {"total": 0, "eqd2_3": 0, "bed_3": 0,
                  "eqd2_10": 0, "bed_10": 0, "eqd2_man": 0, "bed_man": 0}
        for c in [self.c1, self.c2, self.c3]:
            d, n, ab = c.get_values()
            totals["total"]    += d * n
            totals["eqd2_3"]   += calc_eqd2(d, n, 3)
            totals["bed_3"]    += calc_bed(d, n, 3)
            totals["eqd2_10"]  += calc_eqd2(d, n, 10)
            totals["bed_10"]   += calc_bed(d, n, 10)
            totals["eqd2_man"] += calc_eqd2(d, n, ab)
            totals["bed_man"]  += calc_bed(d, n, ab)
        self.sum_total.setText(f"{totals['total']:.2f}")
        self.sum_eqd2_3.setText(f"{totals['eqd2_3']:.2f}")
        self.sum_bed_3.setText(f"{totals['bed_3']:.2f}")
        self.sum_eqd2_10.setText(f"{totals['eqd2_10']:.2f}")
        self.sum_bed_10.setText(f"{totals['bed_10']:.2f}")
        self.sum_eqd2_man.setText(f"{totals['eqd2_man']:.2f}")
        self.sum_bed_man.setText(f"{totals['bed_man']:.2f}")

    # ── OAR table ─────────────────────────────────────────────
    def update_oar_table(self):
        """
        C1_reduced = C1_EQD2 × (1 - c1_red%)
        C2_reduced = C2_EQD2 × (1 - c2_red%)   [only when C3 has values]
        SUM        = C1_reduced + C2_reduced + C3_EQD2
        Allowed    = Dose Limit − SUM
        > 0 → green OK  |  ≤ 0 → red EXCEEDS ALLOWED  |  no limit → yellow Review
        """
        c1_apply = self.c1_red_check.isChecked()
        c1_t_idx = self.c1_time_combo.currentIndex() if c1_apply else -1
        c2_apply = self.c2_red_check.isChecked() and self.c3.has_values()
        c2_t_idx = self.c2_time_combo.currentIndex() if c2_apply else -1

        for row, oar in enumerate(OAR_DATA):
            c1_red_pct = oar["reduction"][c1_t_idx] if c1_apply else 0
            c2_red_pct = oar["reduction"][c2_t_idx] if c2_apply else 0
            limit = oar["limit"]

            red_str = f"{c1_red_pct}%" if limit is not None else "—"
            self.oar_table.item(row, COL_RED).setText(red_str)
            self.oar_table.item(row, COL_RED).setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            red2_str = f"{c2_red_pct}%" if limit is not None else "—"
            self.oar_table.item(row, COL_RED2).setText(red2_str)
            self.oar_table.item(row, COL_RED2).setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            c1_raw, c2_raw, c3_raw = self.oar_raw[row]

            c1_display = c1_raw * (1 - c1_red_pct / 100)
            c2_display = c2_raw * (1 - c2_red_pct / 100)
            c3_display = c3_raw

            for col, val in [(COL_C1, c1_display), (COL_C2, c2_display), (COL_C3, c3_display)]:
                it = self.oar_table.item(row, col)
                it.setText(f"{val:.2f}")
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            total_plan = c1_display + c2_display + c3_display
            sum_it = self.oar_table.item(row, COL_SUM)
            sum_it.setText(f"{total_plan:.2f}")
            sum_it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            allow_it = self.oar_table.item(row, COL_ALLOW)
            allow_it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if limit is not None:
                allowed = limit - total_plan
                allow_it.setText(f"{allowed:.2f}")
                if allowed > 0:
                    allow_it.setBackground(QBrush(QColor(59, 125, 212)))
                    allow_it.setForeground(QBrush(QColor(255, 255, 255)))
                    status = "OK";        color = QColor(100, 220, 100)
                else:
                    allow_it.setBackground(QBrush(QColor(220, 50, 50)))
                    allow_it.setForeground(QBrush(QColor(255, 255, 255)))
                    status = "VIOLATION"; color = QColor(255, 80, 80)
            else:
                allow_it.setBackground(QBrush(QColor(59, 125, 212)))
                allow_it.setForeground(QBrush(QColor(255, 255, 255)))
                allow_it.setText("—")
                status = "REVIEW";        color = QColor(255, 220, 50)

            si = QTableWidgetItem(status)
            si.setBackground(QBrush(color))
            si.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if status == "VIOLATION":
                si.setForeground(QBrush(QColor(255, 255, 255)))
            self.oar_table.setItem(row, COL_STATUS, si)

        self.apply_row_highlight()

    def update_sum_and_oar(self):
        self.update_sum()
        self.update_oar_table()

    # ── Clear ─────────────────────────────────────────────────
    def clear_all(self):
        """Clear courses and non-frozen OAR rows."""
        self.current_oar_idx = None
        self.oar_selector.setCurrentIndex(0)
        self._set_courses_enabled(False)

        for c in [self.c1, self.c2, self.c3]:
            c.clear()
        self.c2_red_check.setEnabled(False)
        self.c2_red_check.setChecked(False)

        for row in range(len(OAR_DATA)):
            if row in self.frozen_oars:
                continue
            self.oar_raw[row] = [0.0, 0.0, 0.0]

        self.update_sum()
        self.update_oar_table()

    # ── Patient info ──────────────────────────────────────────
    def get_patient_info(self):
        return {
            "name":      self.patient_name.text() or "—",
            "id":        self.patient_id.text() or "—",
            "dob":       self.dob.date().toString("dd/MM/yyyy"),
            "diagnosis": self.diagnosis.text() or "—",
            "physician": self.physician.text() or "—",
            "plan_date": self.plan_date.date().toString("dd/MM/yyyy"),
            "notes":     self.notes.text() or "—",
        }

    # ── PDF ───────────────────────────────────────────────────
    def export_pdf(self):
        p = self.get_patient_info()
        default_name = f"{p['id']}.pdf" if p['id'] != "—" else "report.pdf"
        path, _ = QFileDialog.getSaveFileName(self, "Save PDF", default_name, "PDF Files (*.pdf)")
        if not path:
            return
        if not path.endswith(".pdf"):
            path += ".pdf"

        doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                                leftMargin=1*cm, rightMargin=1*cm,
                                topMargin=1.5*cm, bottomMargin=1*cm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('title', fontSize=14, alignment=TA_CENTER, fontName='Helvetica-Bold')
        elements = []

        p = self.get_patient_info()
        elements.append(Paragraph("RE-IRRADIATION DOSE CALCULATION", title_style))
        elements.append(Spacer(1, 0.3*cm))

        pt_data = [["Patient:", p["name"], "ID:", p["id"]],
                   ["C1 Plan Date:", p["dob"], "C2 Plan Date:", p["plan_date"]],
                   ["Medical Physicist:", p["diagnosis"], "Physician:", p["physician"]],
                   ["Notes:", p["notes"], "", ""]]
        pt_t = Table(pt_data, colWidths=[3*cm, 7*cm, 3*cm, 7*cm])
        pt_t.setStyle(TableStyle([
            ('FONTNAME', (0,0),(-1,-1),'Helvetica'),
            ('FONTNAME', (0,0),(0,-1),'Helvetica-Bold'),
            ('FONTNAME', (2,0),(2,-1),'Helvetica-Bold'),
            ('FONTSIZE', (0,0),(-1,-1),9),
            ('GRID', (0,0),(-1,-1),0.3,colors.grey),
            ('BACKGROUND', (0,0),(-1,-1),colors.whitesmoke),
        ]))
        elements.append(pt_t)
        elements.append(Spacer(1, 0.4*cm))

        # Courses
        ch = ["", "Dose/Fr\n(Gy)", "# Fr", "Total\n(Gy)", "EQD2\n(α/β=3)", "BED\n(α/β=3)",
              "EQD2\n(α/β=10)", "BED\n(α/β=10)", "Manual\nα/β", "EQD2\n(manual)", "BED\n(manual)"]
        cr = [ch]
        for lbl, c in [("COURSE 1", self.c1), ("COURSE 2", self.c2), ("COURSE 3", self.c3)]:
            d, n, ab = c.get_values()
            cr.append([lbl, f"{d:.2f}", f"{n:.2f}", f"{d*n:.2f}",
                f"{calc_eqd2(d,n,3):.2f}", f"{calc_bed(d,n,3):.2f}",
                f"{calc_eqd2(d,n,10):.2f}", f"{calc_bed(d,n,10):.2f}",
                f"{ab}", f"{c.get_eqd2_manual():.2f}", f"{calc_bed(d,n,ab):.2f}"])
        cr.append(["SUM", "—", "—", self.sum_total.text(),
            self.sum_eqd2_3.text(), self.sum_bed_3.text(),
            self.sum_eqd2_10.text(), self.sum_bed_10.text(),
            "—", self.sum_eqd2_man.text(), self.sum_bed_man.text()])
        ct = Table(cr, colWidths=[2.2*cm, 2.0*cm, 1.5*cm, 1.8*cm, 2.2*cm, 2.2*cm, 2.2*cm, 2.2*cm, 1.5*cm, 2.4*cm, 2.4*cm])
        ct.setStyle(TableStyle([
            ('BACKGROUND', (0,0),(-1,0),colors.HexColor('#2E4057')),
            ('TEXTCOLOR', (0,0),(-1,0),colors.white),
            ('FONTNAME', (0,0),(-1,0),'Helvetica-Bold'),
            ('FONTNAME', (0,-1),(-1,-1),'Helvetica-Bold'),
            ('BACKGROUND', (0,-1),(-1,-1),colors.HexColor('#dde8f0')),
            ('FONTSIZE', (0,0),(-1,-1),8),
            ('GRID', (0,0),(-1,-1),0.3,colors.grey),
            ('ALIGN', (1,0),(-1,-1),'CENTER'),
            ('VALIGN', (0,0),(-1,-1),'MIDDLE'),
        ]))
        elements.append(Paragraph("Course Summary", styles['Heading2']))
        elements.append(ct)
        elements.append(Spacer(1, 0.4*cm))

        # OAR table
        c1_apply = self.c1_red_check.isChecked()
        c1_t_idx = self.c1_time_combo.currentIndex() if c1_apply else -1
        c2_apply = self.c2_red_check.isChecked() and self.c3.has_values()
        c2_t_idx = self.c2_time_combo.currentIndex() if c2_apply else -1

        oar_hdr = ["Frz", "OAR", "α/β", "Limit\n(EQD2)", "C1\nReduction%", "C2\nReduction%",
                   "Allowed\nEQD2", "C1\nEQD2", "C2\nEQD2", "C3\nEQD2", "SUM", "Status"]
        oar_rows = [oar_hdr]
        for row, oar in enumerate(OAR_DATA):
            c1v = self.oar_table.item(row, COL_C1).text()
            c2v = self.oar_table.item(row, COL_C2).text()
            c3v = self.oar_table.item(row, COL_C3).text()
            has_data = any(v not in ("0", "0.00", "") for v in [c1v, c2v, c3v])
            if not (has_data or row in self.frozen_oars):
                continue
            c1_red = oar["reduction"][c1_t_idx] if c1_apply else 0
            c2_red = oar["reduction"][c2_t_idx] if c2_apply else 0
            limit = oar["limit"]
            limit_str = f"{limit}" if limit is not None else "—"
            allowed_str = self.oar_table.item(row, COL_ALLOW).text()
            oar_rows.append([
                "✓" if row in self.frozen_oars else "",
                oar["name"], str(oar["ab"]), limit_str,
                f"{c1_red}%" if limit is not None else "—",
                f"{c2_red}%" if limit is not None else "—",
                allowed_str, c1v, c2v, c3v,
                self.oar_table.item(row, COL_SUM).text(),
                self.oar_table.item(row, COL_STATUS).text(),
            ])

        if len(oar_rows) > 1:
            ow = [0.8*cm, 4.0*cm, 1.1*cm, 1.5*cm, 1.7*cm, 1.7*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 2.3*cm]
            ot = Table(oar_rows, colWidths=ow)
            os = TableStyle([
                ('BACKGROUND', (0,0),(-1,0),colors.HexColor('#2E4057')),
                ('TEXTCOLOR', (0,0),(-1,0),colors.white),
                ('FONTNAME', (0,0),(-1,0),'Helvetica-Bold'),
                ('FONTSIZE', (0,0),(-1,-1),7.5),
                ('GRID', (0,0),(-1,-1),0.3,colors.grey),
                ('ALIGN', (1,0),(-1,-1),'CENTER'),
                ('ALIGN', (1,0),(1,-1),'LEFT'),
                ('VALIGN', (0,0),(-1,-1),'MIDDLE'),
                ('ROWBACKGROUNDS', (0,1),(-1,-1),[colors.white, colors.HexColor('#f5f5f5')]),
                ('TEXTCOLOR', (6,1),(6,-1), colors.white),
            ])
            sc_map = {"OK": '#64dc64', "VIOLATION": '#ff5050', "REVIEW": '#ffdc32'}
            for ri in range(1, len(oar_rows)):
                if oar_rows[ri][0] == "✓":
                    os.add('BACKGROUND', (0,ri),(5,ri), colors.HexColor('#d2ebff'))
                    os.add('BACKGROUND', (7,ri),(10,ri), colors.HexColor('#d2ebff'))
                # Allowed column (index 6): blue if ≥ 0 or "—", red if < 0
                allowed_val = oar_rows[ri][6]
                try:
                    allow_bg = '#3b7dd4' if float(allowed_val) >= 0 else '#dc3232'
                except ValueError:
                    allow_bg = '#3b7dd4'
                os.add('BACKGROUND', (6,ri),(6,ri), colors.HexColor(allow_bg))
                sc = sc_map.get(oar_rows[ri][-1], '#c8c8c8')
                os.add('BACKGROUND', (11,ri),(11,ri), colors.HexColor(sc))
                if oar_rows[ri][-1] == "VIOLATION":
                    os.add('TEXTCOLOR', (11,ri),(11,ri), colors.white)
            ot.setStyle(os)
            c1_info = f"C1 reduction: {TIME_INTERVALS[c1_t_idx]}" if c1_apply else "No C1 reduction"
            c2_info = f"  |  C2 reduction: {TIME_INTERVALS[c2_t_idx]}" if c2_apply else ""
            elements.append(Paragraph("OAR Dose Limits", styles['Heading2']))
            elements.append(Paragraph(c1_info + c2_info, styles['Normal']))
            elements.append(Spacer(1, 0.2*cm))
            elements.append(ot)

        doc.build(elements)
        QMessageBox.information(self, "Done", f"PDF saved:\n{path}")

    # ── Excel ─────────────────────────────────────────────────
    def export_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save Excel", "", "Excel Files (*.xlsx)")
        if not path:
            return
        if not path.endswith(".xlsx"):
            path += ".xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "RE-IRRADIATION Calculator"
        hf = PatternFill("solid", fgColor="2E4057")
        hfont = Font(bold=True, color="FFFFFF", size=10)
        bf = Font(bold=True, size=10)
        ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
        th = Side(style="thin")
        bd = Border(left=th, right=th, top=th, bottom=th)

        p = self.get_patient_info()
        row = 1
        ws.cell(row,1,"LQ Model Calculator — Radiotherapy Dose Planner").font = Font(bold=True,size=13)
        row += 2
        for lbl, val in [("Patient",p["name"]),("ID",p["id"]),("C1 Plan Date",p["dob"]),
                         ("C2 Plan Date",p["plan_date"]),("Medical Physicist",p["diagnosis"]),
                         ("Physician",p["physician"]),("Notes",p["notes"])]:
            ws.cell(row,1,lbl).font=bf; ws.cell(row,2,val); row+=1
        row += 1

        for col,h in enumerate(["Course","Dose/Fr (Gy)","# Fractions","Total Dose (Gy)",
                                 "EQD2 (α/β=3)","BED (α/β=3)","EQD2 (α/β=10)","BED (α/β=10)",
                                 "EQD2 (manual)","BED (manual)"],1):
            c=ws.cell(row,col,h); c.fill=hf; c.font=hfont; c.alignment=ctr; c.border=bd
        row += 1
        for lbl,course in [("COURSE 1",self.c1),("COURSE 2",self.c2),("COURSE 3",self.c3)]:
            d,n,ab=course.get_values()
            for col,v in enumerate([lbl,d,round(n,2),round(d*n,2),
                round(calc_eqd2(d,n,3),2),round(calc_bed(d,n,3),2),
                round(calc_eqd2(d,n,10),2),round(calc_bed(d,n,10),2),
                round(course.get_eqd2_manual(),2),round(calc_bed(d,n,ab),2)],1):
                c=ws.cell(row,col,v); c.border=bd; c.alignment=ctr
            row+=1
        sum_fill=PatternFill("solid",fgColor="dde8f0")
        for col,v in enumerate(["SUM","—","—",self.sum_total.text(),self.sum_eqd2_3.text(),
                self.sum_bed_3.text(),self.sum_eqd2_10.text(),self.sum_bed_10.text(),
                self.sum_eqd2_man.text(),self.sum_bed_man.text()],1):
            c=ws.cell(row,col,v); c.font=bf; c.fill=sum_fill; c.border=bd; c.alignment=ctr
        row += 2

        for col,h in enumerate(["Frozen","OAR","α/β","Dose Limit (EQD2 Gy)","C1 Red%","C2 Red%",
                                 "Allowed Remained EQD2 (Gy)","C1 Plan EQD2","C2 Plan EQD2",
                                 "C3 Plan EQD2","SUM PLAN EQD2","Status"],1):
            c=ws.cell(row,col,h); c.fill=hf; c.font=hfont; c.alignment=ctr; c.border=bd
        row += 1

        c1_apply = self.c1_red_check.isChecked()
        c1_t_idx = self.c1_time_combo.currentIndex() if c1_apply else -1
        c2_apply = self.c2_red_check.isChecked() and self.c3.has_values()
        c2_t_idx = self.c2_time_combo.currentIndex() if c2_apply else -1
        sc_xl = {"OK":"64dc64","VIOLATION":"ff5050","REVIEW":"ffdc32"}
        frz_fill = PatternFill("solid",fgColor="d2ebff")
        wfont = Font(color="FFFFFF", size=10)
        for oi,oar in enumerate(OAR_DATA):
            c1_red = oar["reduction"][c1_t_idx] if c1_apply else 0
            c2_red = oar["reduction"][c2_t_idx] if c2_apply else 0
            limit=oar["limit"]; limit_str=limit if limit is not None else "—"
            is_frz=oi in self.frozen_oars
            status=self.oar_table.item(oi,COL_STATUS).text()
            allow_str=self.oar_table.item(oi,COL_ALLOW).text()
            try:
                allow_bg = "3b7dd4" if float(allow_str) >= 0 else "dc3232"
            except ValueError:
                allow_bg = "3b7dd4"
            vals=["✓" if is_frz else "",oar["name"],oar["ab"],limit_str,
                  f"{c1_red}%" if limit is not None else "—",
                  f"{c2_red}%" if limit is not None else "—",
                  allow_str,
                  self.oar_table.item(oi,COL_C1).text(),
                  self.oar_table.item(oi,COL_C2).text(),
                  self.oar_table.item(oi,COL_C3).text(),
                  self.oar_table.item(oi,COL_SUM).text(),status]
            alt=PatternFill("solid",fgColor="f5f5f5") if oi%2 else None
            for col,v in enumerate(vals,1):
                c=ws.cell(row,col,v); c.border=bd; c.alignment=ctr
                if col==12: c.fill=PatternFill("solid",fgColor=sc_xl.get(status,"c8c8c8"))
                elif col==7:
                    c.fill=PatternFill("solid",fgColor=allow_bg); c.font=wfont
                elif is_frz: c.fill=frz_fill
                elif alt: c.fill=alt
            row+=1

        ws.column_dimensions['A'].width=8; ws.column_dimensions['B'].width=30
        for lt in ['C','D','E','F','G','H','I','J','K']:
            ws.column_dimensions[lt].width=16
        wb.save(path)
        QMessageBox.information(self,"Done",f"Excel saved:\n{path}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    win = MainWindow()
    win.show()
    sys.exit(app.exec())
